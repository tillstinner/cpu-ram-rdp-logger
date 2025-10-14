import psutil
import time
import subprocess
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# -----------------------------
LOG_FILE = "cpu_ram_rdp_log.xlsx"
INTERVAL = 2
CPU_THRESHOLD = 10
RAM_THRESHOLD = 10
# -----------------------------

def is_rdp_active():
    try:
        result = subprocess.run(["query", "session"], capture_output=True, text=True)
        for line in result.stdout.splitlines():
            if "rdp-tcp" in line.lower() and "aktiv" in line.lower():
                return True
        return False
    except Exception as e:
        print(f"Error checking RDP session: {e}")
        return False

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

def collect_data_xlsx(filename, interval=2, cpu_threshold=10, ram_threshold=10, laufzeit=0):
    """Collect CPU, RAM, RDP usage, write tick data to Excel, print & write summary."""
    total_time = 0
    time_over_cpu = 0
    time_over_ram = 0
    time_over_rdp = 0
    last_time = time.time()

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Daten"
    ws_data.append(["Zeit", "CPU_%", "RAM_%", "RDP_active"])

    start_time = time.time()
    try:
        while True:
            now = time.time()
            if laufzeit > 0 and now - start_time > laufzeit:
                break

            cpu = psutil.cpu_percent(interval=None)
            ram = psutil.virtual_memory().percent
            rdp_status = "Ja" if is_rdp_active() else "Nein"
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            ws_data.append([timestamp, round(cpu,1), round(ram,1), rdp_status])

            # Console output per tick
            print(f"{timestamp} -> CPU: {cpu:.1f}%, RAM: {ram:.1f}%, RDP: {rdp_status}")

            # Update counters
            delta = now - last_time
            total_time += delta
            if cpu > cpu_threshold: time_over_cpu += delta
            if ram > ram_threshold: time_over_ram += delta
            if rdp_status == "Ja": time_over_rdp += delta

            last_time = now
            time.sleep(interval)

    except KeyboardInterrupt:
        print("\nLogging stopped by user.")

    # Function to format and write/print summary
    def write_and_print_summary(ws_summary):
        anteil_cpu = (time_over_cpu / total_time * 100) if total_time else 0
        anteil_ram = (time_over_ram / total_time * 100) if total_time else 0
        anteil_rdp = (time_over_rdp / total_time * 100) if total_time else 0

        ws_summary.append(["Metrik", "Wert", "Anteil (%)"])
        ws_summary.append(["Gesamtdauer (s)", round(total_time,1), 100])
        ws_summary.append([f"CPU ueber {cpu_threshold}%", round(time_over_cpu,1), round(anteil_cpu,1)])
        ws_summary.append([f"RAM ueber {ram_threshold}%", round(time_over_ram,1), round(anteil_ram,1)])
        ws_summary.append(["RDP aktiv", round(time_over_rdp,1), round(anteil_rdp,1)])

        # Also print to terminal
        print("\n--- Zusammenfassung ---")
        print(f"Gesamtdauer: {total_time:.1f} s")
        print(f"CPU ueber {cpu_threshold}%: {time_over_cpu:.1f} s ({anteil_cpu:.1f}%)")
        print(f"RAM ueber {ram_threshold}%: {time_over_ram:.1f} s ({anteil_ram:.1f}%)")
        print(f"RDP aktiv: {time_over_rdp:.1f} s ({anteil_rdp:.1f}%)")

    # Create summary sheet
    ws_summary = wb.create_sheet(title="Zusammenfassung")
    write_and_print_summary(ws_summary)

    # Auto-adjust widths
    auto_adjust_column_width(ws_data)
    auto_adjust_column_width(ws_summary)

    wb.save(filename)
    print(f"\nExcel file saved: {filename}")

def main():
    print("CPU & RAM Logger with RDP status")
    try:
        laufzeit = float(input("Enter duration in seconds (0 = infinite): ").strip())
    except ValueError:
        laufzeit = 0

    collect_data_xlsx(LOG_FILE, INTERVAL, CPU_THRESHOLD, RAM_THRESHOLD, laufzeit)

if __name__ == "__main__":
    main()
