import psutil
import time
import subprocess
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt # type: ignore
import os


# -----------------------------
INTERVAL = 2
CPU_THRESHOLD = 10
RAM_THRESHOLD = 10
# -----------------------------


def is_rdp_active():
    """Check if any RDP session is currently active"""
    try:
        result = subprocess.run(["query", "session"], capture_output=True, text=True)
        for line in result.stdout.splitlines():
            if "rdp-tcp" in line.lower() and "aktiv" in line.lower():
                return True
        return False
    except Exception as e:
        print(f"Error checking RDP session: {e}")
        return False


def auto_adjust_column_width(ws):
    """Automatically adjust column widths based on cell content"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2


def create_or_load_workbook(filename):
    """Load existing workbook or create new one"""
    if os.path.exists(filename):
        try:
            wb = load_workbook(filename)
            if "Daten" in wb.sheetnames:
                ws_data = wb["Daten"]
            else:
                ws_data = wb.create_sheet("Daten")
                ws_data.append(["Zeit", "CPU_%", "RAM_%", "RDP_active", "RDP_numeric"])
        except Exception:
            print(f"Datei '{filename}' beschädigt – wird neu erstellt.")
            os.remove(filename)
            wb = Workbook()
            ws_data = wb.active
            ws_data.title = "Daten"
            ws_data.append(["Zeit", "CPU_%", "RAM_%", "RDP_active", "RDP_numeric"])
    else:
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Daten"
        ws_data.append(["Zeit", "CPU_%", "RAM_%", "RDP_active", "RDP_numeric"])
    return wb, ws_data


# =============================================
#   DATENERFASSUNG
# =============================================

def collect_data_xlsx(filename, interval=2, laufzeit=0):
    """Collect data and save continuously to Excel"""
    wb, ws_data = create_or_load_workbook(filename)

    start_time = time.time()
    try:
        while True:
            now = time.time()
            if laufzeit > 0 and now - start_time > laufzeit:
                break

            cpu = psutil.cpu_percent(interval=None)
            ram = psutil.virtual_memory().percent
            rdp_status = "Ja" if is_rdp_active() else "Nein"
            rdp_numeric = 1 if rdp_status == "Ja" else 0
            timestamp = datetime.now()

            ws_data.append([
                timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                round(cpu, 1),
                round(ram, 1),
                rdp_status,
                rdp_numeric
            ])
            wb.save(filename)

            os.system('cls' if os.name == 'nt' else 'clear')
            print(f"{timestamp.strftime('%Y-%m-%d %H:%M:%S')} -> CPU: {cpu:.1f}%, RAM: {ram:.1f}%, RDP: {rdp_status}")
            print("")
            print("Logging aktiv. Fenster bitte NICHT schliessen. Berechnungen können durchgeführt werden")
            time.sleep(interval)

    except KeyboardInterrupt:
        print("\nLogging stopped by user.")

    wb.save(filename)
    print(f"Data saved to {filename}")


# =============================================
#   AUSWERTUNG & VISUALISIERUNG
# =============================================

def analyze_and_visualize(filename, chart_file, cpu_threshold=10, ram_threshold=10):
    """Read data, compute stats, create summary + chart"""

    # --- Load workbook and data ---
    wb = load_workbook(filename)
    ws_data = wb["Daten"]
    data_rows = list(ws_data.iter_rows(min_row=2, values_only=True))

    if not data_rows:
        print("Keine Daten vorhanden für Auswertung.")
        return

    timestamps, cpu_list, ram_list, rdp_list = [], [], [], []

    for row in data_rows:
        timestamps.append(datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S"))
        cpu_list.append(float(row[1]))
        ram_list.append(float(row[2]))
        rdp_list.append(int(row[4]))

    # --- Statistik ---
    total_time = (timestamps[-1] - timestamps[0]).total_seconds() if len(timestamps) > 1 else 0
    time_over_cpu = 0
    time_over_ram = 0
    time_over_rdp = 0

    for i in range(1, len(timestamps)):
        delta = (timestamps[i] - timestamps[i-1]).total_seconds()
        if cpu_list[i] > cpu_threshold:
            time_over_cpu += delta
        if ram_list[i] > ram_threshold:
            time_over_ram += delta
        if rdp_list[i] == 1:
            time_over_rdp += delta

    anteil_cpu = (time_over_cpu / total_time * 100) if total_time else 0
    anteil_ram = (time_over_ram / total_time * 100) if total_time else 0
    anteil_rdp = (time_over_rdp / total_time * 100) if total_time else 0

    if "Zusammenfassung" in wb.sheetnames:
        wb.remove(wb["Zusammenfassung"])
    ws_summary = wb.create_sheet(title="Zusammenfassung")

    ws_summary.append(["Metrik", "Wert", "Anteil (%)"])
    ws_summary.append(["Gesamtdauer (s)", round(total_time, 1), 100])
    ws_summary.append([f"CPU über {cpu_threshold}%", round(time_over_cpu, 1), round(anteil_cpu, 1)])
    ws_summary.append([f"RAM über {ram_threshold}%", round(time_over_ram, 1), round(anteil_ram, 1)])
    ws_summary.append(["RDP aktiv", round(time_over_rdp, 1), round(anteil_rdp, 1)])

    # --- Matplotlib Chart ---
    plt.figure(figsize=(12, 6))
    rdp_scaled = [x * 100 for x in rdp_list]

    plt.plot(timestamps, cpu_list, label="CPU %", color="blue")
    plt.plot(timestamps, ram_list, label="RAM %", color="green")
    for i in range(len(rdp_scaled) - 1):
        plt.hlines(rdp_scaled[i], timestamps[i], timestamps[i + 1],
                   color="red", label="RDP aktiv (0/1)" if i == 0 else "")

    plt.xlabel("Zeit")
    plt.ylabel("CPU/RAM (%) / RDP (0/1)")
    plt.title("CPU, RAM & RDP Verlauf")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(chart_file)
    plt.close()

    # --- Add chart image to summary ---
    img = XLImage(chart_file)
    img.anchor = "E2"
    ws_summary.add_image(img)

    # --- Adjust column widths ---
    for ws in [ws_data, ws_summary]:
        auto_adjust_column_width(ws)

    wb.save(filename)

    print("\n--- Zusammenfassung ---")
    print(f"Gesamtdauer: {total_time:.1f} s")
    print(f"CPU über {cpu_threshold}%: {time_over_cpu:.1f} s ({anteil_cpu:.1f}%)")
    print(f"RAM über {ram_threshold}%: {time_over_ram:.1f} s ({anteil_ram:.1f}%)")
    print(f"RDP aktiv: {time_over_rdp:.1f} s ({anteil_rdp:.1f}%)")
    print(f"Chart eingebettet in: {filename}")


# =============================================
#   MAIN-FUNKTION
# =============================================

def main():
    print("CPU & RAM Logger with RDP status")
    print("--------------------------------")

    try:
        laufzeit = float(input("Enter duration in seconds (0 = infinite): ").strip())
    except ValueError:
        laufzeit = 0

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    LOG_FILE = f"cpu_ram_rdp_log_{timestamp}.xlsx"
    CHART_FILE = f"cpu_ram_rdp_chart_{timestamp}.png"

    # Daten sammeln
    collect_data_xlsx(LOG_FILE, INTERVAL, laufzeit)

    # Auswertung automatisch danach
    analyze_and_visualize(LOG_FILE, CHART_FILE, CPU_THRESHOLD, RAM_THRESHOLD)


if __name__ == "__main__":
    main()
